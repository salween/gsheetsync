# -*- coding: utf-8 -*-
"""Main module."""

import glob
import hashlib
from datetime import datetime
from lxml import etree
from lxml import objectify
import unicodecsv as csv
import openpyxl

Element = objectify.Element
E = objectify.E

def field_to_header(field, attribute):
    return '{0}:{1}'.format(field.tag, attribute)

def get_header(content, ct):
    """Get header for a content type, by crawling all the content items
    """
    header_dict = {}
    for item in content.findall(ct):
        for field in item:
            for attribute in field.keys():
                header_dict[field_to_header(field, attribute)] = None
    return sorted(header_dict.keys())

def csv_from_contenttype(content, ct):
    """
    Create a CSV file for an Orchard ContentType.

    Given a child of an Orchard <Types> element, generate a worksheet with
    columns for the fields.
    """
    header = get_header(content, ct)

    with open('{0}.csv'.format(ct), 'wb') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header)
        writer.writeheader()

        for item in content.findall(ct):
            i = item.get('Id')
            if 'IdentityPart:Identifier' in header:
                row = {'IdentityPart:Identifier': i.split('=')[1]}
            else:
                row = {}
            for field in item:
                for attribute, value in field.items():
                    row[field_to_header(field, attribute)] = value
            writer.writerow(row)

def csvs_from_contenttypes(et):
    types = et.find('.//Types')
    content = et.find('.//Content')
    if types:
        contenttypes = [ct.tag for ct in types]
    else:
        contenttypes = set([ct.tag for ct in content])
    for ct in contenttypes:
        csv_from_contenttype(content, ct)

def xlsx_from_csvs():
    csv_filenames = glob.glob("*.csv")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for fn in csv_filenames:
        ct = fn.replace('.csv', '')
        ws = wb.create_sheet(ct)
        with open(fn, 'rb') as csvfile:
            data = csv.reader(csvfile)
            for row in data:
                ws.append(row)
    wb.save('orchard.xlsx')
    return wb

def xlsx_from_xml(infile):
    et = etree.fromstring(infile.read())
    # parts = et.find('.//Parts')

    csvs_from_contenttypes(et)
    xlsx_from_csvs()


def csv_from_xlsx(infile):
    wb = openpyxl.load_workbook(infile.name)
    for sheetname in wb.sheetnames:
        ws = wb.get_sheet_by_name(sheetname)
        with open('{0}.csv'.format(sheetname), 'wb') as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows():
                writer.writerow([c.value for c in row])

def sha_id(row):
    keys = list(row.keys())
    keys.sort()
    values = [row[k] for k in keys]
    sha1 = hashlib.sha1()
    sha1.update(repr(values).encode('utf-8'))
    return sha1.hexdigest()

def xml_from_csv():
    nowstr = datetime.now().strftime('%Y-%m-%d')
    orchard = Element('Orchard')
    orchard.Recipe = E.Recipe(E.ExportUtc(nowstr))
    # orchard.Feature = E.Feature(enable=', '.join(site_description[0]['Features']))
    # orchard.Settings = E.Settings(*list(genSettings(site_description[1]['Settings'])))
    orchard.Migration = E.Migration(features="*")
    # orchard.Command = E.Command("""
    #     layer create Default /LayerRule:"true" /Description:"The widgets in this layer are displayed on all pages"
    #     ...
    #     """)

    csv_filenames = glob.glob("*.csv")
    orchard.Content = E.Content()
    for fn in csv_filenames:
        ct = fn.replace('.csv', '')
        with open(fn, 'rb') as csvfile:
            data = csv.DictReader(csvfile)
            for row in data:
                item = {}
                for fieldattr, value in row.items():
                    tagname, attribute = fieldattr.split(':')
                    if value:
                        attr_values = item.setdefault(tagname, {})
                        attr_values[attribute] = value
                if not item.get('IdentityPart'):
                    item['IdentityPart'] = {'Identifier': sha_id(row)}
                el = getattr(E, ct)(
                    *[getattr(E, tag)(attributes) for tag, attributes in item.items()],
                    Id="/Identifier=%s"%item['IdentityPart']['Identifier'],
                    Status="Published"
                )
                orchard.Content.append(el)

    objectify.deannotate(orchard, cleanup_namespaces=True)
    ff = open('import.xml', 'wb')
    ff.write(etree.tostring(orchard, pretty_print=True))
    ff.close()

def xml_from_xlsx(infile):
    csv_from_xlsx(infile)
    xml_from_csv()
